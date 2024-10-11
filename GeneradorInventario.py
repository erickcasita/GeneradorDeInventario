from Utils import headers,title_category,styles_conten_category,styles_totales
from Connection import connection
from openpyxl import load_workbook
import time,datetime,os,shutil,locale,threading
from progress1bar import ProgressBar
from helpers import validatedate,sendMail
from Activate import Get_current_date, Get_current_expiration_software, Check_expiration_software


print("""
    ░█████╗░░█████╗░███╗░░░███╗███████╗██████╗░  ░██████╗░█████╗░████████╗
    ██╔══██╗██╔══██╗████╗░████║██╔════╝██╔══██╗  ██╔════╝██╔══██╗╚══██╔══╝
    ██║░░╚═╝██║░░██║██╔████╔██║█████╗░░██████╔╝  ╚█████╗░███████║░░░██║░░░
    ██║░░██╗██║░░██║██║╚██╔╝██║██╔══╝░░██╔══██╗  ░╚═══██╗██╔══██║░░░██║░░░
    ╚█████╔╝╚█████╔╝██║░╚═╝░██║███████╗██║░░██║  ██████╔╝██║░░██║░░░██║░░░
    ░╚════╝░░╚════╝░╚═╝░░░░░╚═╝╚══════╝╚═╝░░╚═╝  ╚═════╝░╚═╝░░╚═╝░░░╚═╝░░░
      """)
print("            GENERADOR DE EXISTENCIAS ALMACENES LLENO SAT Y JDC           ")
print("                             VERSIÓN 10                                  ")
print("                         ING. ERICK CASANOVA                             ")
print("                           ÁREA DE SISTEMAS                              ")
print("   |-------------------------------------------------------------------|")
print("   |------------------ 1.- CREAR INVENTARIO  --------------------------|")
print("   |------------------ 2.- ENVIAR INVENTARIO --------------------------|")
print("   |------------------ 3.-    SALIR          --------------------------|")
print("   |-------------------------------------------------------------------|")

print("\n ....... Validando versión y fecha de expiración ....... ")
date_software = Get_current_expiration_software()
date_server = Get_current_date()
flag = False;
response =  Check_expiration_software(date_server, date_software)
if(response == True):
    flag = True;
elif(response == False):
    print("\n ... Error en la licencía de expiración ...  || Contacte a su administrador ")
    flag = False;
    input("\n Ingrese cualquier tecla para salir ");
else:
    print(response);
    flag = False;
while flag:
    
    option = input("\n Ingrese una opción: ")
    try:
        option = int(option)
    except ValueError:
       print ('\n Ingrese una opción válida, Por favor')
       continue
    if(option == 1):
        date = input("\n Ingrese Fecha cierre de Almacén YYY-mm-dd: ")
        if(validatedate(date)):
            print("\n ........ Creando Inventario en excel ..... \n")
            #diccionary category beer
            data = {
            1: "CERVEZA",
            2: "REFRESCOS",
            3: "AGUA NATURAL",
            10: "JUGOS",
            11: "BEBIDAS ENERGÉTICAS",
            14: "AGUA MINERAL",
            17: "BEBIDAS ISOTÓNICAS",
            18: "ADAS",
            22: "COMESTIBLES"
            
        }
            #array totals category
            totals_category = []
            headers()
            #iteration dictorary category
            fila = 4  #row initial
            #Call file inventario.xlsx
            wb = load_workbook("inventario.xlsx")
            ws = wb.active
            for key in data:
                #Insert title category
                title_category(wb,fila,str(data[key]))
                fila=fila+1
                #Conection Sql Server
                con = connection()
                sql = con.cursor()
                #Execute procedure
                command = "set nocount on; execute GeneradorDeInventarioXCategoriasV2 @clavecategoria=?, @fechacierre=?"
                params = (key,str(date))
                sql.execute(command,params)
                rows = sql.fetchall()
                #Get total rows
                totalprodcts = len(rows)
                #kwargs  config progressBar
                kwargs = {
                    'total': totalprodcts,
                    'completed_message': 'Categoria ' +data[key]+" completa",
                    'clear_alias': True,
                    'show_fraction': False,
                    'show_prefix': False,
                    'show_duration': True
                    }
                with ProgressBar(**kwargs) as pb:
                    for row in rows:
          
                        ws.cell(fila,1).value = row[0]
                        sku = row[1]
                        sku = str(sku).strip()
                        if(sku == ""):
                            sku = "0"
                        ws.cell(fila,2).value = int(sku)
                        ws.cell(fila,3).value = row[2]
                        ws.cell(fila,4).value = row[3]
                        ws.cell(fila,5).value = row[4]
                        ws.cell(fila,6).value = row[5]
                        ws.cell(fila,11).value = row[6]
                        ws.cell(fila,12).value = row[7]
                        ws.cell(fila,13).value = row[8]
                        ws.cell(fila,18).value = row[9]
                        ws.cell(fila,19).value = row[10]
                        ws.cell(fila,20).value = row[11]
                        #Apply Styles conten category
                        styles_conten_category(ws,fila)
                        #Oculting rows 
                        if(row[11] == 0):
                            ws.row_dimensions[fila].hidden = True
                        fila = fila+1
                        #ProgressBar
                        pb.alias = row[2]
                        time.sleep(.08)
                        pb.count += 1
                #Add to array position total x category
                totals_category.append(fila+1)
                #Add Total x category
                ws.cell(fila+1,3).value = "TOTAL"
                ws.cell(fila+1,6).value = "=SUM(F"+str(fila-totalprodcts)+":F"+str(fila-1)+")"
                ws.cell(fila+1,13).value = "=SUM(M"+str(fila-totalprodcts)+":M"+str(fila-1)+")"
                ws.cell(fila+1,20).value = "=SUM(T"+str(fila-totalprodcts)+":T"+str(fila-1)+")"
                #Styles totales
                styles_totales(ws,fila+1)
                fila=fila+3
                sql.close()  
                #wb.save("inventario.xlsx") 
            ws.cell(fila+1,3).value = "GRAN TOTAL"
            ws.cell(fila+1,6).value = "=SUM(F"+str(totals_category[0])+", F"+str(totals_category[1])+",F"+str(totals_category[2])+",F"+str(totals_category[3])+",F"+str(totals_category[4])+",F"+str(totals_category[5])+",F"+str(totals_category[6])+",F"+str(totals_category[7])+",F"+str(totals_category[8])+")"
            ws.cell(fila+1,13).value = "=SUM(M"+str(totals_category[0])+", M"+str(totals_category[1])+",M"+str(totals_category[2])+",M"+str(totals_category[3])+",M"+str(totals_category[4])+",M"+str(totals_category[5])+",M"+str(totals_category[6])+",M"+str(totals_category[7])+",M"+str(totals_category[8])+")"
            ws.cell(fila+1,20).value = "=SUM(T"+str(totals_category[0])+", T"+str(totals_category[1])+",T"+str(totals_category[2])+",T"+str(totals_category[3])+",T"+str(totals_category[4])+",T"+str(totals_category[5])+",T"+str(totals_category[6])+",T"+str(totals_category[7])+",T"+str(totals_category[8])+")"
            #Styles Gran total
            styles_totales(ws,fila+1)
            wb.save("inventario.xlsx")
            #Administrator file
            locale.setlocale(locale.LC_ALL, 'es_ES.utf8')
            #namefile = 'INVENTARIO DE ALMACENES AL DIA '+str(datetime.datetime.strftime(datetime.datetime.now(),'%d-%m-%Y %H-%M-%S %p'))+'.xlsx'
            namefile = "Inventario de almacenes al dia " + str(datetime.datetime.strftime(datetime.datetime.now(),'%A %d de %B del %Y %H %M %S %p'))+ '.xlsx'
            savepath = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Documents','ReporteadorInventario',namefile)
            savepath = os.path.abspath(savepath) 
            shutil.copy('inventario.xlsx', savepath)
            os.remove('inventario.xlsx')
            #Creating file name attachment mail
            with open("mails/attachment.name.mail", 'w+') as fichero:
                    fichero.writelines(namefile)
            print("\n ........Inventario en excel terminado ..... ")           
    if(option == 2):
        print("\n")
        if(os.path.isfile("mails/attachment.name.mail")):
            date = datetime.datetime.strftime(datetime.datetime.now(),'%d/%m/%Y')
            remitente =   "almacensat@coronalostuxtlas.com.mx"
            destinatario = ["direcciongral@coronalostuxtlas.com.mx"]
            work_send_mail = threading.Thread(name="send_email", target=sendMail)
            work_send_mail.start()
            work_send_mail.join()
            print("\n ................ Envío Terminado ..................")
        else:
            print("\n El Inventario ya ha sido enviado o no ha generado uno nuevo ")
    if(option == 3):
        flag = False;
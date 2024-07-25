from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, numbers
import datetime,locale

def title_category(file,row,categoryname):
    wb = load_workbook(file)
    ws = wb.active
    ws.cell(row,1).value = "CLAVE INT"
    ws.cell(row,1).font = Font(color="000000", size="11", name="Arial", bold=True)
    ws.cell(row,1).alignment = Alignment(horizontal='center',vertical="center")
    ws.cell(row,2).value = "SKU"
    ws.cell(row,2).font = Font(color="000000", size="11", name="Arial", bold=True)
    ws.cell(row,2).alignment = Alignment(horizontal='center',vertical="center")
    ws.cell(row,3).value = categoryname
    ws.cell(row,3).font = Font(color="000000", size="11", name="Arial", bold=True)
    ws.cell(row,3).alignment = Alignment(horizontal='center',vertical="center")
    ws.cell(row,4).value = "Tarimas"
    ws.cell(row,4).font = Font(color="000000", size="12", name="Arial")
    ws.cell(row,4).alignment = Alignment(horizontal='center',vertical="center")
    ws.cell(row,5).value = "Saldos"
    ws.cell(row,5).font = Font(color="000000", size="12", name="Arial")
    ws.cell(row,5).alignment = Alignment(horizontal='center',vertical="center")
    ws.cell(row,6).value = "Unidades"
    ws.cell(row,6).font = Font(color="000000", size="12", name="Arial")
    ws.cell(row,6).alignment = Alignment(horizontal='center',vertical="center")
    ws.cell(row,11).value = "Tarimas"
    ws.cell(row,11).font = Font(color="000000", size="12", name="Arial")
    ws.cell(row,11).alignment = Alignment(horizontal='center',vertical="center")
    ws.cell(row,12).value = "Saldos"
    ws.cell(row,12).font = Font(color="000000", size="12", name="Arial")
    ws.cell(row,12).alignment = Alignment(horizontal='center',vertical="center")
    ws.cell(row,13).value = "Unidades"
    ws.cell(row,13).font = Font(color="000000", size="12", name="Arial")
    ws.cell(row,13).alignment = Alignment(horizontal='center',vertical="center")
    ws.cell(row,18).value = "Tarimas"
    ws.cell(row,18).font = Font(color="000000", size="12", name="Arial")
    ws.cell(row,18).alignment = Alignment(horizontal='center',vertical="center")
    ws.cell(row,19).value = "Saldos"
    ws.cell(row,19).font = Font(color="000000", size="12", name="Arial")
    ws.cell(row,19).alignment = Alignment(horizontal='center',vertical="center")
    ws.cell(row,20).value = "Unidades"
    ws.cell(row,20).font = Font(color="000000", size="12", name="Arial")
    ws.cell(row,20).alignment = Alignment(horizontal='center',vertical="center")
    wb.save("inventario.xlsx")
def headers():
    wb = Workbook()
    ws = wb.active
    ws.sheet_view.showGridLines = False
    ws['B2'] = "Reporte de Inventario Físico de Líquido"
    ws['B2'].font = Font(color="000000", size="14", name="Comic Sans MS", underline="single")
    ws['B2'].alignment = Alignment(horizontal='general',vertical='center')
    ws.row_dimensions[2].height = 47.40
    #Unions Cells for date
    ws.merge_cells('H2:K2')
    locale.setlocale(locale.LC_ALL, 'es_ES.utf8')
    date = datetime.datetime.strftime(datetime.datetime.now(),'%A, %d de %B del %Y')
    ws['H2'] = date
    ws['H2'].number_format = numbers.FORMAT_DATE_TIME5
    ws['H2'].font = Font(color="000000", size="12", name="Arial", bold=True)
    ws['H2'].alignment = Alignment(horizontal='center')
    #Unions Cell for  headers "Almacén"
    ws.merge_cells('B3:D3')
    ws['B3'] = "ALMACÉN"
    ws['B3'].font = Font(color="FFFFFF", size="13.5", name="MS Sans Serif")
    ws['B3'].alignment = Alignment(horizontal='center',vertical="center")
    #ws['B3'].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
    ws.row_dimensions[3].height = 47.40
    ws.column_dimensions['C'].width = 53.67
    #Unions Cell for Headers "San Andres" and "Juan Diaz Covarrubias" and "Totales"
    ws.merge_cells('E3:G3')
    ws.merge_cells('K3:N3')
    ws.merge_cells('R3:T3')
    ws['E3'] = "San Andrés"
    ws['E3'].font = Font(color="FFFFFF", size="11", name="Comic Sans MS")
    ws['E3'].alignment = Alignment(horizontal='center',vertical="center")
    ws['K3'] = "Juan Díaz Covarrubias"
    ws['K3'].font = Font(color="FFFFFF", size="11", name="Comic Sans MS")
    ws['K3'].alignment = Alignment(horizontal='center',vertical="center")
    ws['R3'] = "TOTALES"
    ws['R3'].font = Font(color="FFFFFF", size="11", name="Comic Sans MS")
    ws['R3'].alignment = Alignment(horizontal='center',vertical="center")
    #Header "Stock Min and Stock Max SAT and JDC"
    ws['I3'] = "Stock Mín"
    ws['J3'] = "Stock MAX"
    ws['P3'] = "Stock Mín"
    ws['Q3'] = "Stock MAX"
    ws['I3'].font = Font(color="FFFF00", size="10", name="Comic Sans MS")
    ws['I3'].alignment = Alignment(wrap_text=True,horizontal='center',vertical="center")
    ws['J3'].font = Font(color="FFFF00", size="10", name="Comic Sans MS")
    ws['J3'].alignment = Alignment(wrap_text=True,horizontal='center',vertical="center")
    ws['P3'].font = Font(color="FFFF00", size="10", name="Comic Sans MS")
    ws['P3'].alignment = Alignment(wrap_text=True,horizontal='center',vertical="center")
    ws['Q3'].font = Font(color="FFFF00", size="10", name="Comic Sans MS")
    ws['Q3'].alignment = Alignment(wrap_text=True,horizontal='center',vertical="center")
    for i in  range(1, 21):
        ws.cell(3,i).fill = PatternFill(fgColor="222b35", fill_type = "solid")
        
    #Rows and columns dimensions  all categories
    ws.row_dimensions[4].height = 18
    ws.column_dimensions['A'].width = 11.89
    ws.column_dimensions['B'].width = 11.22
    ws.column_dimensions['D'].width = 10.67
    ws.column_dimensions['E'].width = 8.89
    ws.column_dimensions['F'].width = 12.11
    ws.column_dimensions['K'].width = 12.78
    ws.column_dimensions['L'].width = 10.67
    ws.column_dimensions['M'].width = 12.22
    ws.column_dimensions['R'].width = 10.67
    ws.column_dimensions['S'].width = 10.67
    ws.column_dimensions['T'].width = 11.89
    wb.save("inventario.xlsx")
    #Insert title category "cerveza"
    title_category("inventario.xlsx",4,"CERVEZA")